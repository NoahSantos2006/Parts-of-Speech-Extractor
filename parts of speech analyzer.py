import nltk
from nltk import word_tokenize
import openpyxl

def POSanalyzer(user_input):
    def retry():
        while True:
            first_try = True
            if first_try:
                new_input = input("Type the following below to continue:\n\nNew text: Type: 'new sentence'.\nStop talking: Type: 'bye'.\nSame sentence: Type:'yes'.\n\nInput: ")
                print()
            else:
                new_input = "That wasn't any of the options silly! Try again: "

            if new_input.lower() == 'new sentence':
                new_input1 = input("Hello again! Send a text and I'll analyze the parts of speech in it! ")
                POSanalyzer(new_input1)
            elif new_input.lower() == 'yes':
                print(f"Here is your sentence: {user_input}")
                POSanalyzer(user_input)
            elif new_input.lower() == 'bye':
                print("It was nice talking to you!\n")
                exit()
            else:
                first_try = False
    def checkList(user_input, checkList1):
        list1 = []
        for value in checkList1:
            if str(user_input).lower() in value:
                list1.append(value)
        if len(list1) == 0:
            return False
        if len(list1) == 1:
            return list1[0]
        else:
            new_list = []
            while new_list == []:
                print(f"Which {user_input} did you mean? A ", end="")
                for value in list1:
                    if list1[-1] == value:
                        print(f"or a {value.strip()}? ", end="")
                    else:
                        print(f"{value.strip()}, ", end="")
                new_input = input()
                print()
                for value1 in list1:
                    if new_input == value1:
                        return new_input
                    elif new_input in value1:
                        new_list.append(str(value1))
        return checkList(new_input, new_list)
    def output(answer, user_input, partofSpeech, partofSpeechsymbol):
        if answer >= 1:
            print(f'You\'re text was: {user_input}\n\nYou used {answer} {partofSpeech.strip()}(s) also known as {partofSpeechsymbol}. Here are the words you used: {posList}')
            answer = 0
            retry()
        else:
            answer = 0
            print(f'You\'re text was: {user_input}\n\nYou didn\'t use a {partofSpeech.strip()} also known as {partofSpeechsymbol}\n')
            retry()

    global pos
    global posValues
    answer = 0
    while True:
        posList = []
        taggedWords = dict(nltk.pos_tag(word_tokenize(user_input)))
        print(taggedWords)
        print('\nHere is a list of the parts of speech:', pos, '\n')
        input1 = input("Which parts of speech would you like to know is in your text? ")
        print()
        if 'bye' in input1:
            print("It was nice talking to you!\n")
            exit()
        elif input1.upper() in pos.keys():
            for a, b in pos.items():
                if a.lower() == input1.lower():
                    partofSpeech = b
                    partofSpeechsymbol = a
                    for c, d in taggedWords.items():
                        if d.lower() == input1.lower():
                            posList.append(c)
                            answer += 1
            output(answer, user_input, partofSpeech, partofSpeechsymbol)
        else:
            userPOS = checkList(input1, posValues)
            if userPOS:
                for a, b in pos.items():
                    if b == userPOS:
                        print(b)
                        partofSpeech = b
                        partofSpeechsymbol = a
                        for c, d in taggedWords.items():
                            if d.lower() == a.lower():
                                posList.append(c)
                                answer += 1
                output(answer, user_input, partofSpeech, partofSpeechsymbol)
            else:
                print("I was not able to find the part of speech you inputted. Please try again.")

        
wb = openpyxl.load_workbook('posSheet.xlsx')
ws = wb.active
pos = {}
for i in range(1, 27):
    pos[ws[f'A{i}'].value] = ws[f'B{i}'].value

posValues = []
for key, value in pos.items():
    posValues.append(value)

user_input = input("Hello! Send a text and I'll analyze the parts of speech in it! ")

POSanalyzer(user_input)